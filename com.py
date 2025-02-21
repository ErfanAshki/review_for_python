from openpyxl import load_workbook

workbook = load_workbook(filename='new_excel.xlsx')
sheet = workbook.active

for i in range(1, sheet.max_column +1):
    cell = sheet.cell(row=i, column=2)
    print(cell.value, end='   ')

print()

for i in range(1, sheet.max_column +1):
    cell = sheet.cell(row=2, column=i)
    print(cell.value, end='   ')

